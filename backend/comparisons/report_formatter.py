
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ExcelReportFormatter:
    """Handles the formatting of Excel reports"""
    
    @staticmethod
    def apply_formatting(report_path):
        """Apply formatting to the Excel report for better readability."""
        try:
            workbook = openpyxl.load_workbook(report_path)
            
            # Format each worksheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Format headers
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Format data cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in sheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                        cell.border = thin_border
                        
                        # Highlight winning documents with light green background
                        if (sheet_name == 'Overall Results' and cell.column_letter == 'D') or sheet_name == 'Win Summary':
                            if (sheet_name == 'Overall Results' and cell.row > 1) or (sheet_name == 'Win Summary' and cell.row > 1):
                                row_cells = list(row)
                                win_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                                for r_cell in row_cells:
                                    r_cell.fill = win_fill
                
                # Adjust column widths
                ExcelReportFormatter._adjust_column_widths(sheet)
                
                # Color code score cells based on their values
                if sheet_name in ['Detailed Criterion Analysis', 'Criterion Scores']:
                    ExcelReportFormatter._apply_score_colors(sheet)
            
            # Add a legend for the scoring colors
            ExcelReportFormatter._add_scoring_legend(workbook)
            
            # Save the formatted workbook
            workbook.save(report_path)
            
        except Exception as e:
            print(f"Warning: Could not format Excel file - {e}")
        finally:
            print("Excel formatting process completed")
    
    @staticmethod
    def _adjust_column_widths(sheet):
        """Adjust column widths based on content"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Find the maximum length in the column
            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            # Limit width to reasonable size and ensure minimum width
            adjusted_width = min(max_length + 2, 50)
            if adjusted_width < 10:
                adjusted_width = 10
            
            # Apply width
            sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Special handling for analysis columns in detailed criterion sheet
            header = column[0].value
            if header and ('Analysis' in header or 'Reasoning' in header):
                sheet.column_dimensions[column_letter].width = 60
    
    @staticmethod
    def _apply_score_colors(sheet):
        """Apply color coding to score cells based on their values"""
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                header = sheet.cell(row=1, column=cell.column).value
                if header and ('Score' in header) and isinstance(cell.value, (int, float)):
                    score_value = float(cell.value)
                    # Color gradient from red (1) to yellow (3) to green (5)
                    if score_value <= 1:
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif score_value <= 2:
                        cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid") 
                    elif score_value <= 3:
                        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    elif score_value <= 4:
                        cell.fill = PatternFill(start_color="E5FFCC", end_color="E5FFCC", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    
    @staticmethod
    def _add_scoring_legend(workbook):
        """Add a legend explaining the color coding for scores"""
        legend_sheet = workbook["Win Summary"] if "Win Summary" in workbook.sheetnames else workbook.create_sheet("Legend")
        legend_row = 2 if "Legend" in workbook.sheetnames else legend_sheet.max_row + 3
        
        legend_sheet.cell(row=legend_row, column=1).value = "Scoring Legend:"
        legend_sheet.cell(row=legend_row, column=1).font = Font(bold=True)
        
        # Add color legend
        colors = [
            ("1 - Poor", "FFCCCC"),
            ("2 - Fair", "FFE5CC"),
            ("3 - Good", "FFFFCC"),
            ("4 - Very Good", "E5FFCC"),
            ("5 - Excellent", "CCFFCC")
        ]
        
        for i, (label, color) in enumerate(colors):
            legend_sheet.cell(row=legend_row + i + 1, column=1).value = label
            legend_sheet.cell(row=legend_row + i + 1, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            legend_sheet.cell(row=legend_row + i + 1, column=1).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
